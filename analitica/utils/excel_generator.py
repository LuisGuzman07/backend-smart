"""
Generador de reportes en formato Excel usando openpyxl
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from django.utils import timezone


def generar_excel(titulo, encabezados, datos, hoja_nombre="Reporte"):
    """
    Genera un archivo Excel con formato básico
    
    Args:
        titulo: Título del reporte
        encabezados: Lista de nombres de columnas
        datos: Lista de listas con los datos
        hoja_nombre: Nombre de la hoja
    
    Returns:
        BytesIO con el contenido del Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = hoja_nombre
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    titulo_font = Font(bold=True, size=14)
    titulo_alignment = Alignment(horizontal="center", vertical="center")
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(encabezados))
    cell_titulo = ws.cell(row=1, column=1)
    cell_titulo.value = titulo
    cell_titulo.font = titulo_font
    cell_titulo.alignment = titulo_alignment
    
    # Fecha de generación
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(encabezados))
    cell_fecha = ws.cell(row=2, column=1)
    cell_fecha.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    cell_fecha.alignment = Alignment(horizontal="right")
    
    # Encabezados
    fila_header = 4
    for col_idx, encabezado in enumerate(encabezados, start=1):
        cell = ws.cell(row=fila_header, column=col_idx)
        cell.value = encabezado
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Datos
    fila_inicio_datos = fila_header + 1
    for row_idx, fila_datos in enumerate(datos, start=fila_inicio_datos):
        for col_idx, valor in enumerate(fila_datos, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Convertir datetime con timezone a naive para Excel
            if isinstance(valor, datetime) and valor.tzinfo is not None:
                valor = timezone.localtime(valor).replace(tzinfo=None)
            
            cell.value = valor
            cell.border = border_style
            
            # Alineación de números a la derecha
            if isinstance(valor, (int, float)):
                cell.alignment = Alignment(horizontal="right")
    
    # Ajustar ancho de columnas
    for col_idx in range(1, len(encabezados) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Guardar en BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file
